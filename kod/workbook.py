from openpyxl import Workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font


class ExcelWriter:
    def __init__(self, col, columns, metadata):
        self.wb = Workbook()
        self.ws = self.wb.active

        size = len(columns)

        # Yıl Okul row
        self.ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=size + 4)

        # Ders Not tipi row
        self.ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=size + 4)

        # Sınıf label
        self.ws.merge_cells(start_row=3, start_column=1, end_row=5, end_column=1)

        # Sınıf ad
        self.ws.merge_cells(start_row=3, start_column=2, end_row=4, end_column=3)

        # Sınıf
        self.ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=3)

        # Kriter label
        self.ws.merge_cells(start_row=3, start_column=4, end_row=4, end_column=size + 3)

        # Kriter
        self.ws.merge_cells(start_row=6, start_column=1, end_row=14, end_column=1)
        for i in range(size):
            self.ws.merge_cells(
                start_row=5, start_column=4 + i, end_row=15, end_column=4 + i
            )
            self.ws.cell(5, 4 + i).alignment = Alignment(
                text_rotation=90, vertical="center", horizontal="center", wrap_text=True
            )
            self.ws.cell(5, 4 + i).font = Font(size=10)
            self.ws.cell(5, 4 + i).value = columns[i]

        self.ws.merge_cells(
            start_row=3, start_column=size + 4, end_row=15, end_column=size + 4
        )

        self.ws.cell(3, size + 4).alignment = Alignment(
            text_rotation=90, vertical="center", horizontal="center"
        )

        for i in range(5):
            self.ws["B" + str(i + 8)] = i + 1

        self.ws["C8"] = "ZAYIF"
        self.ws["C9"] = "KABUL EDİLEBİLİR"
        self.ws["C10"] = "ORTA"
        self.ws["C11"] = "İYİ"
        self.ws["C12"] = "ÇOK İYİ"

        self.ws["A15"] = "SIRA"
        self.ws["A15"].font = Font(bold=True)
        self.ws["B15"] = "NO"
        self.ws["B15"].font = Font(bold=True)
        self.ws["C15"] = "ADI SOYADI"
        self.ws["C15"].font = Font(bold=True)

        self.ws["A1"].alignment = Alignment(horizontal="center")
        self.ws["A2"].alignment = Alignment(horizontal="center")

        self.ws["A3"].alignment = Alignment(
            text_rotation=90, vertical="center", horizontal="center"
        )
        self.ws["A3"] = "SINIF"
        self.ws["A3"].font = Font(size=16, bold=True)

        self.ws["A6"] = "ÖLÇÜTLER"
        self.ws["A6"].alignment = Alignment(
            text_rotation=90, vertical="center", horizontal="center"
        )
        self.ws["A6"].font = Font(size=18, bold=True)

        self.ws["D3"].alignment = Alignment(horizontal="center")
        self.ws["D3"] = "Öğrencide Gözlenecek kazanımlar"
        self.ws["D3"].font = Font(size=18, bold=True)

        self.ws["B3"].alignment = Alignment(horizontal="center")
        self.ws["B3"].font = Font(size=18, bold=True)
        self.ws["B3"] = metadata[3]

        self.ws["B4"].alignment = Alignment(horizontal="center")
        self.ws["B5"].alignment = Alignment(horizontal="center")
        self.ws["B5"] = "SINIFI"
        self.ws["b5"].font = Font(bold=True)

        self.ws["A1"] = metadata[5] + " EĞİTİM ÖĞRETİM YILI " + metadata[0]
        self.ws["A2"] = (
            metadata[4]
            + " DERSİ "
            + metadata[6]
            + " "
            + col.upper()
            + " NOTU DEĞERLENDİRME ÖLÇEĞİ"
        )
        self.ws["A1"].font = Font(bold=True)
        self.ws["A2"].font = Font(bold=True)

        self.ws.cell(3, size + 4).value = col.upper() + " PUANI"
        self.ws.cell(3, size + 4).font = Font(size=18, bold=True)

    def add_borders(self, style):
        border = Border(
            left=Side(style=style),
            right=Side(style=style),
            top=Side(style=style),
            bottom=Side(style=style),
        )

        for row in self.ws.iter_rows():
            for cell in row:
                cell.border = border

    def add_dataframe(self, df):
        data = df.values.tolist()
        for row in data:
            self.ws.append(row)
        self.add_borders("thin")
        self.fit_names()

    def fit_names(self):
        names = []
        for cell in self.ws["C"]:
            value = cell.value
            if value != None:
                names.append(cell.value)
        max_length = len(max(names, key=len))
        adjusted_width = (max_length + 2) * 1.2
        self.ws.column_dimensions["C"].width = adjusted_width

    def add_names(self, teacher, principal, subject):
        max_r = self.ws.max_row
        max_c = self.ws.max_column
        self.ws.cell(max_r + 5, 3).value = teacher
        self.ws.cell(max_r + 6, 3).value = subject + " ÖĞRETMENİ"
        self.ws.cell(max_r + 6, 3).font = Font(bold=True)

        self.ws.cell(max_r + 5, max_c - 3).value = principal
        self.ws.cell(max_r + 6, max_c - 3).value = "OKUL MÜDÜRÜ"
        self.ws.cell(max_r + 6, max_c - 3).font = Font(bold=True)

    def save(self, filename="taslak.xlsx"):
        self.wb.save(filename)
